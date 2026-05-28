#!/usr/bin/env python3
"""Build the SuperGlue performance-model Excel FROM SCRATCH (one script, no
incremental patches), in the MobileViT sheet's exact style, and SELF-VALIDATE:
for every produced sheet we recompute the grand-total cycles independently in
pure Python and compare against the spreadsheet's formula-evaluated result.

Sheets produced (kept alongside the source MobileViT):
  SuperGlue (original)             log-domain Sinkhorn, dual-softmax readout
  SG max-Sinkhorn                  Sinkhorn -> sum + scale  (our sinkhorn_iter/col)
  SG base2-softmax                 attn softmax via exp2    (same cost)
  SG transpose-free col            col norm uses cross-row max, no transpose
  SG fixed-thresh match            readout = argmax + threshold (1 pass)
  SG max-Sinkhorn (measured)       Cycles measured by running each kernel in the
                                   IPU emulator (matmul / softmax / sinkhorn).

Bare-op model: one elementary vector pass per row (multiply, accumulate, max,
sum, exp2, sub, log, move, compare) with cycles = elements/128. A MAC is TWO
rows (multiply + accumulate). Time = Cycles/freq[GHz]/1000.

Run: python build_sg_excel.py [src.xlsx] [out.xlsx]
"""
import os, sys, re, math
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter

# ---------------- parameters -------------------------------------------------
N, D, HEADS, HD, NP1, T = 512, 256, 4, 64, 513, 100
GNN_BLOCKS = 18 * 2                                            # 18 self/cross x 2 images
IMG_H, IMG_W = 480, 640                                        # SuperPoint input

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    "/root/.claude/uploads/eecfeba7-80e7-4f8a-8032-ec4e983b04a4/cb06a587-IPU_2.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/home/user/ipu-emulator/IPU_2_SuperGlue.xlsx"

# ---------------- bare-op row recipe -----------------------------------------
# (group, op_name, c, d, contraction K, kind, instance_multiplier)
#   kind 'mac'  -> S = C*D*G/128   (one of multiply/accumulate; emit 2 rows for full MAC)
#   kind 'ew'   -> S = C*D/128     (elementwise pass: max/sum/exp2/sub/log/mul/add/move/cmp)
#   kind 'rsh'  -> S = 0           (reshape/view; T=0)
def mac(group, c, d, K):
    return [(group, "multiply", c, d, K, "mac"), (group, "accumulate", c, d, K, "mac")]


def rows_for_superpoint(K_kpts=N):
    """Per-image SuperPoint rows (encoder + detector + descriptor). Multiplied by
    2 in the grand total (one image per side of the pair)."""
    rows = []
    # ---- shared encoder: 4 stages of (conv,conv) + 2x2/s2 maxpool ----
    enc = [(IMG_H, IMG_W, [(1, 64, 'conv1a'), (64, 64, 'conv1b')]),
           (IMG_H // 2, IMG_W // 2, [(64, 64, 'conv2a'), (64, 64, 'conv2b')]),
           (IMG_H // 4, IMG_W // 4, [(64, 128, 'conv3a'), (128, 128, 'conv3b')]),
           (IMG_H // 8, IMG_W // 8, [(128, 128, 'conv4a'), (128, 128, 'conv4b')])]
    cur_C = 0
    for i, (H, W, convs) in enumerate(enc):
        for cin, cout, name in convs:
            rows += mac(f'{name} 3x3+ReLU', cout, H * W, cin * 9)
            cur_C = cout
        if i < len(enc) - 1:                                   # pool between stages (3 pools)
            outH, outW = H // 2, W // 2
            rows.append((f'maxpool 2x2/s2 -> {outH}x{outW}', 'max (4 taps)',
                         cur_C, outH * outW, 4, 'mac'))         # 4 shifted-max taps
    # ---- detector head (at 60x80) ----
    H, W = IMG_H // 8, IMG_W // 8
    rows += mac('convPa 3x3+ReLU', 256, H * W, 128 * 9)
    rows += mac('convPb 1x1 (->65)', 65, H * W, 256)
    for op in ['max', 'subtract', 'exp2', 'sum', 'multiply(1/Z)']:
        rows.append(('detector softmax(65)', op, 65, H * W, 0, 'ew'))
    rows.append(('depth-to-space', 'reshape', 64, H * W, 0, 'rsh'))
    Hf, Wf = H * 8, W * 8                                       # 480 x 640 heatmap
    rows.append(('simple_nms 9x9 iter 3', 'maxpool 243 taps', Hf, Wf, 243, 'mac'))
    rows.append(('score threshold', 'relu(s-tau)', Hf, Wf, 0, 'ew'))
    rows.append(('topk cap (calibrated)', 'soft count + bisect', Hf, Wf, 30, 'mac'))
    # ---- descriptor head (at 60x80) ----
    rows += mac('convDa 3x3+ReLU', 256, H * W, 128 * 9)
    rows += mac('convDb 1x1 (->256)', 256, H * W, 256)
    for op in ['mul (squares)', 'sum-reduce', 'multiply (scale by rsqrt)']:
        rows.append(('descriptor L2 (dense)', op, 256, H * W, 0, 'ew'))
    rows += mac('grid_sample (4-corner)', 256, K_kpts, 4)
    for op in ['mul (squares)', 'sum-reduce', 'multiply (scale by rsqrt)']:
        rows.append(('per-keypoint L2', op, 256, K_kpts, 0, 'ew'))
    return rows


def rows_for_variant(variant):
    """Return the full list of (group, op, c, d, K, kind) rows for one SG variant
    (no multipliers baked in: per-row instance counts are folded into the GNN/
    head/Sinkhorn TOTAL formulas, exactly like MobileViT's '4X heads' / '2 blocks')."""
    rows = []
    # ---- Keypoint encoder (per image; x2 via Z totals) ----
    rows.append(("Input", "kpt+desc load", D, N, 0, "rsh"))
    for cin, cout in [(3, 32), (32, 64), (64, 128), (128, 256)]:
        rows += mac(f"KptEnc {cin}->{cout}", cout, N, cin)
    rows.append(("Residual desc+=enc", "add", D, N, 0, "ew"))
    # ---- One GNN attention block (heads expanded x4 below) ----
    for nm in "QKV":
        rows += mac(f"{nm} proj D->D", D, N, D)
    for nm in "QKV":
        rows.append((f"head reshape {nm}", "reshape", HD, N, 0, "rsh"))
    # head core (later x4): QKt + scale + softmax (5 ew) + attn@V
    rows += mac("QKt (1 head)", N, N, HD)
    rows.append(("scale 1/sqrt(d)", "multiply", N, N, 0, "ew"))
    for op in ["max", "subtract", "exp2", "sum", "multiply(1/Z)"]:
        rows.append(("softmax (1 head)", op, N, N, 0, "ew"))
    rows += mac("attn @ V (1 head)", N, HD, N)
    # block tail
    rows.append(("concat heads", "reshape", D, N, 0, "rsh"))
    rows += mac("out proj D->D", D, N, D)
    rows.append(("Residual x+=msg", "add", D, N, 0, "ew"))
    rows += mac("merge MLP 2D->2D", 2 * D, N, 2 * D)
    rows.append(("merge MLP relu", "relu", 2 * D, N, 0, "ew"))
    rows += mac("merge MLP 2D->D", D, N, 2 * D)
    rows.append(("Residual x+=mlp", "add", D, N, 0, "ew"))
    # ---- Matching ----
    rows += mac("final proj D->D", D, N, D)                    # x2 images via total
    rows += mac("score = A^T B", N, N, D)
    rows.append(("scale 1/sqrt(d)", "multiply", N, N, 0, "ew"))
    rows.append(("dustbin augment", "reshape", NP1, NP1, 0, "rsh"))
    # Sinkhorn -- variant differences
    if variant == "max_sinkhorn":
        row_ops = ["sum", "multiply(1/r)"]; col_ops = ["sum", "multiply(1/c)"]
    else:
        row_ops = ["max", "exp2", "sum", "log", "subtract"]
        col_ops = ["max", "exp2", "sum", "log", "subtract"]
    for op in row_ops:
        rows.append(("Sinkhorn row norm", op, NP1, NP1, 0, "ew"))
    if variant == "transpose_free":
        rows.append(("Sinkhorn col transpose (eliminated)", "reshape", NP1, NP1, 0, "rsh"))
    else:
        rows.append(("Sinkhorn col transpose", "move", NP1, NP1, 0, "ew"))
    for op in col_ops:
        rows.append(("Sinkhorn col norm", op, NP1, NP1, 0, "ew"))
    # Match readout
    if variant == "fixed_thresh":
        for op in ["max(argmax)", "compare(threshold)"]:
            rows.append(("match readout", op, NP1, NP1, 0, "ew"))
    else:
        for op in ["max", "exp2", "sum", "subtract", "multiply"]:
            rows.append(("match readout (dual-softmax)", op, NP1, NP1, 0, "ew"))
    return rows


# ---------------- spreadsheet plumbing --------------------------------------
COLS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
wb = openpyxl.load_workbook(SRC)
mob = wb["MobileViT"]
HEADER = [mob.cell(1, c).value for c in range(1, 27)]


def cstyle(dst, src):
    dst.font = copy(src.font); dst.fill = copy(src.fill); dst.border = copy(src.border)
    dst.alignment = copy(src.alignment); dst.number_format = src.number_format


def make_sheet(name):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for c in range(1, 27):
        ws.cell(1, c, HEADER[c - 1]); cstyle(ws.cell(1, c), mob.cell(1, c))
    for i in range(1, 27):
        L = get_column_letter(i); ws.column_dimensions[L].width = mob.column_dimensions[L].width
    ws.freeze_panes = "B2"
    ws["Y2"], ws["Z2"] = 1, 8
    return ws


def emit_row(ws, r, group, op, c, d, g, kind):
    S = {"mac": f"=C{r}*D{r}*G{r}/128", "ew": f"=C{r}*D{r}/128", "rsh": "0"}[kind]
    vals = {"A": group, "B": op, "C": c, "D": d, "E": 0, "F": 0, "G": g, "H": 0,
            "I": 0, "J": 0, "K": 0, "L": c, "M": d, "N": 0,
            "O": f"=C{r}*D{r}/1024", "P": 0, "Q": 0, "R": f"=L{r}*M{r}/1024", "S": S}
    if kind == "rsh":
        vals["S"] = "0"; vals["T"] = 0
        for cc in "UVWX":
            vals[cc] = 0
    else:
        vals["T"] = f"=(S{r}/$Y$2)/1000"
        for cc, src in [("U", "O"), ("V", "P"), ("W", "Q"), ("X", "R")]:
            vals[cc] = f"=(({src}{r}*8*1024)/$T{r})/1000"
    for col, v in vals.items():
        cstyle(ws.cell(r, COLS.index(col) + 1, v), mob.cell(5, COLS.index(col) + 1))


def write_total(ws, r, label, zformula):
    cstyle(ws.cell(r, 25, label), mob.cell(68, 25))
    cstyle(ws.cell(r, 26, zformula), mob.cell(68, 26))


def build_sheet(variant, name):
    ws = make_sheet(name)
    sp_rows = rows_for_superpoint(K_kpts=N)
    sg_rows = rows_for_variant(variant)
    rows = sp_rows + sg_rows
    sp_end = 1 + len(sp_rows)                                  # last SP row (1-indexed)
    # find indices for the per-image / per-head / per-block / x-T totals
    idx_groups = {}
    for i, (g, *_) in enumerate(rows):
        idx_groups.setdefault(g, []).append(2 + i)             # row 2 onwards
    for i, row in enumerate(rows):
        emit_row(ws, 2 + i, *row)
    last = 1 + len(rows)
    # SuperPoint per-image total (multiplied by 2 in grand total)
    write_total(ws, sp_end, "SuperPoint per image [us] ->",
                f"=SUM(T2:T{sp_end})")
    # totals (encoder x2, head core x4 baked into block, GNN block x36, Sinkhorn x100)
    enc_rows = [r for r in range(2, last + 1) if ws.cell(r, 1).value
                and ("KptEnc" in str(ws.cell(r, 1).value) or "Residual desc" in str(ws.cell(r, 1).value))]
    head_rows = [r for r in range(2, last + 1)
                 if ws.cell(r, 1).value in ("QKt (1 head)", "scale 1/sqrt(d)", "softmax (1 head)", "attn @ V (1 head)")
                 and r < idx_groups.get("concat heads", [last + 1])[0]]
    # the GNN block is everything from "Q proj D->D" through "Residual x+=mlp"
    block_start = idx_groups["Q proj D->D"][0]
    block_end = idx_groups["Residual x+=mlp"][-1]
    sink_rows = sorted(set(idx_groups.get("Sinkhorn row norm", [])
                            + idx_groups.get("Sinkhorn col norm", [])
                            + idx_groups.get("Sinkhorn col transpose", [])))
    read_rows = sorted(set(idx_groups.get("match readout", [])
                            + idx_groups.get("match readout (dual-softmax)", [])))
    final_rows = idx_groups["final proj D->D"]
    score_rows = idx_groups["score = A^T B"]
    scale_match = [r for r in idx_groups["scale 1/sqrt(d)"] if r not in head_rows][0]

    # encoder x2
    write_total(ws, enc_rows[-1], "Encoder total [us] ->",
                f"=2*(SUM(T{enc_rows[0]}:T{enc_rows[-1]}))")
    # 4 heads total
    write_total(ws, idx_groups["concat heads"][0], "total 4 heads [us]",
                f"=4*(SUM(T{head_rows[0]}:T{head_rows[-1]}))")
    # GNN block (sum of T's in block + extra 3*head-sum so heads x4 total, not x1)
    gnn_z = (f"={GNN_BLOCKS}*(SUM(T{block_start}:T{block_end})"
             f"+3*(SUM(T{head_rows[0]}:T{head_rows[-1]})))")
    write_total(ws, block_end, "GNN total [us] ->", gnn_z)
    # Sinkhorn x T
    write_total(ws, sink_rows[-1], f"Sinkhorn total (x{T}) [us] ->",
                f"={T}*(SUM(T{sink_rows[0]}:T{sink_rows[-1]}))")
    # grand total: 2*SuperPoint (per image x 2) + 2*encoder + GNN + 2*final + score
    #              + match scale + Sinkhorn + readout
    gt = last + 2
    cstyle(ws.cell(gt, 1, "GRAND TOTAL [us]"), mob.cell(5, 1))
    grand = (f"=2*Z{sp_end}+Z{enc_rows[-1]}+Z{block_end}"
             f"+2*(SUM(T{final_rows[0]}:T{final_rows[-1]}))"
             f"+SUM(T{score_rows[0]}:T{score_rows[-1]})+T{scale_match}"
             f"+Z{sink_rows[-1]}+SUM(T{read_rows[0]}:T{read_rows[-1]})")
    write_total(ws, gt, "variant: " + variant, grand)
    # FPS
    fr = gt + 1
    cstyle(ws.cell(fr, 1, "FPS (image pair)"), mob.cell(5, 1))
    write_total(ws, fr, "frames / sec ->", f"=1000000/Z{gt}")
    return ws, gt


# ---------------- measured sheet (runs the actual kernels in the emulator) ---
def build_measured_sheet():
    """SG max-Sinkhorn with Cycles MEASURED by running the kernels."""
    sys.path.insert(0, os.path.join(os.path.dirname(__file__),
                                    "kernels/superpoint_superglue/superpoint_eval"))
    from measure_kernels import rates
    R = rates()
    def tiles(e): return math.ceil(e / 128)
    def opcyc(kind, c, d, K):
        if kind == "mac":
            return tiles(c * d) * (R["mac_step"] * K + R["tile_ovh"])
        if kind == "ew":     return tiles(c * d) * R["ew_tile"]
        if kind == "taps":   return tiles(c * d) * K * R["ew_tile"]
        if kind == "sm":     return d * tiles(c) * R["softmax128"]
        if kind == "sr":     return d * tiles(c) * R["sink_row"]
        if kind == "sc":     return d * tiles(c) * R["sink_col"]
        return 0.0
    ROWS = []
    # --- SuperPoint (per image; mult=2 for the pair) ---
    enc_stages = [(IMG_H, IMG_W, [(1, 64, 'conv1a'), (64, 64, 'conv1b')]),
                  (IMG_H // 2, IMG_W // 2, [(64, 64, 'conv2a'), (64, 64, 'conv2b')]),
                  (IMG_H // 4, IMG_W // 4, [(64, 128, 'conv3a'), (128, 128, 'conv3b')]),
                  (IMG_H // 8, IMG_W // 8, [(128, 128, 'conv4a'), (128, 128, 'conv4b')])]
    for i, (H, W, cs) in enumerate(enc_stages):
        for cin, cout, name in cs:
            ROWS.append((f'{name} 3x3+ReLU', 'MAC', cout, H * W, cin * 9, 'mac', 2))
        if i < len(enc_stages) - 1:
            outH, outW = H // 2, W // 2
            ROWS.append((f'maxpool 2x2/s2 -> {outH}x{outW}', 'max(4 taps)',
                         cs[-1][1], outH * outW, 4, 'taps', 2))
    Hd, Wd = IMG_H // 8, IMG_W // 8
    ROWS += [
        ('convPa 3x3+ReLU', 'MAC', 256, Hd * Wd, 128 * 9, 'mac', 2),
        ('convPb 1x1 (->65)', 'MAC', 65, Hd * Wd, 256, 'mac', 2),
        ('detector softmax(65)', 'softmax', 65, Hd * Wd, 0, 'sm', 2),
        ('depth-to-space', 'reshape', 64, Hd * Wd, 0, 'rsh', 2),
        ('simple_nms 9x9 iter 3', 'maxpool', IMG_H, IMG_W, 243, 'taps', 2),
        ('score threshold', 'relu(s-tau)', IMG_H, IMG_W, 0, 'ew', 2),
        ('topk cap (calibrated)', 'soft count + bisect', IMG_H, IMG_W, 30, 'taps', 2),
        ('convDa 3x3+ReLU', 'MAC', 256, Hd * Wd, 128 * 9, 'mac', 2),
        ('convDb 1x1 (->256)', 'MAC', 256, Hd * Wd, 256, 'mac', 2),
        ('descriptor L2 (dense)', '3 passes', 256, Hd * Wd, 3, 'taps', 2),
        ('grid_sample (4-corner)', 'MAC', 256, N, 4, 'mac', 2),
        ('per-keypoint L2', '3 passes', 256, N, 3, 'taps', 2),
    ]
    # --- SuperGlue (max-Sinkhorn variant) ---
    ROWS += [("Input", "kpt+desc load", D, N, 0, "rsh", 1)]
    for cin, cout in [(3, 32), (32, 64), (64, 128), (128, 256)]:
        ROWS.append((f"KptEnc {cin}->{cout}", "MAC matmul", cout, N, cin, "mac", 2))
    ROWS.append(("Residual desc+=enc", "add", D, N, 0, "ew", 2))
    g = GNN_BLOCKS
    ROWS += [
        ("Q proj D->D", "MAC", D, N, D, "mac", g),
        ("K proj D->D", "MAC", D, N, D, "mac", g),
        ("V proj D->D", "MAC", D, N, D, "mac", g),
        ("QKt", "MAC", N, N, HD, "mac", g * HEADS),
        ("scale", "multiply", N, N, 0, "ew", g * HEADS),
        ("softmax", "max,exp2,sum,mul", N, N, 0, "sm", g * HEADS),
        ("attn @ V", "MAC", N, HD, N, "mac", g * HEADS),
        ("out proj", "MAC", D, N, D, "mac", g),
        ("Residual x+=msg", "add", D, N, 0, "ew", g),
        ("merge MLP 2D->2D", "MAC", 2 * D, N, 2 * D, "mac", g),
        ("merge MLP relu", "relu", 2 * D, N, 0, "ew", g),
        ("merge MLP 2D->D", "MAC", D, N, 2 * D, "mac", g),
        ("Residual x+=mlp", "add", D, N, 0, "ew", g),
        ("final proj D->D", "MAC", D, N, D, "mac", 2),
        ("score = A^T B", "MAC", N, N, D, "mac", 1),
        ("scale", "multiply", N, N, 0, "ew", 1),
        ("Sinkhorn row norm (max)", "sum+scale", NP1, NP1, 0, "sr", T),
        ("Sinkhorn col norm (max)", "ACC.MAX+scale", NP1, NP1, 0, "sc", T),
        ("match readout", "argmax+thresh", NP1, NP1, 0, "ew", 1),
    ]
    ws = make_sheet("SG max-Sinkhorn (measured)")
    total = 0.0
    for i, (gr, op, c, d, K, kind, mult) in enumerate(ROWS):
        r = 2 + i
        cyc = round(opcyc(kind, c, d, K) * mult)
        total += cyc
        vals = {"A": gr, "B": op, "C": c, "D": d, "E": K, "F": 0, "G": K, "H": mult,
                "I": 0, "J": 0, "K": 0, "L": c, "M": d, "N": 0,
                "O": f"=C{r}*D{r}/1024", "P": 0, "Q": 0, "R": f"=L{r}*M{r}/1024",
                "S": cyc, "T": f"=(S{r}/$Y$2)/1000"}
        if kind == "rsh" or cyc == 0:
            vals["S"] = 0; vals["T"] = 0
            for cc in "UVWX": vals[cc] = 0
        else:
            for cc, src in [("U", "O"), ("V", "P"), ("W", "Q"), ("X", "R")]:
                vals[cc] = f"=(({src}{r}*8*1024)/$T{r})/1000"
        for col, v in vals.items():
            cstyle(ws.cell(r, COLS.index(col) + 1, v), mob.cell(5, COLS.index(col) + 1))
        if mult > 1:
            ws.cell(r, 27, f"x{mult}")                # col AA -- don't overwrite Y2 (freq)
    gt = 2 + len(ROWS) + 1
    cstyle(ws.cell(gt, 1, "GRAND TOTAL [us] (measured)"), mob.cell(5, 1))
    write_total(ws, gt, "measured emulator cycles ->", f"=SUM(T2:T{1 + len(ROWS)})")
    fr = gt + 1
    cstyle(ws.cell(fr, 1, "FPS (image pair)"), mob.cell(5, 1))
    write_total(ws, fr, "frames / sec ->", f"=1000000/Z{gt}")
    return ws, gt, total, R


# ---------------- Python-side recomputation (validation oracle) -------------
def py_total(variant):
    us = lambda cyc: cyc / 1000.0
    mac_us = lambda c, d, K: 2 * us(c * d * K / 128.0)          # mul + add
    ew_us = lambda c, d: us(c * d / 128.0)
    # SuperPoint per image (x2 for the pair)
    sp = 0
    # ops emitted in the sheet as a SINGLE row with S=c*d*g/128 (one pass per tap)
    # cost 1x, not 2x (these are max/threshold/sigmoid taps, not MACs)
    taps_us = lambda c, d, K: K * ew_us(c, d)
    enc_stages = [(IMG_H, IMG_W, [(1, 64), (64, 64)]),
                  (IMG_H // 2, IMG_W // 2, [(64, 64), (64, 64)]),
                  (IMG_H // 4, IMG_W // 4, [(64, 128), (128, 128)]),
                  (IMG_H // 8, IMG_W // 8, [(128, 128), (128, 128)])]
    for i, (H, W, cs) in enumerate(enc_stages):
        for cin, cout in cs:
            sp += mac_us(cout, H * W, cin * 9)
        if i < len(enc_stages) - 1:
            sp += taps_us(cs[-1][1], (H // 2) * (W // 2), 4)     # 2x2/s2 maxpool: 4 taps
    H, W = IMG_H // 8, IMG_W // 8
    sp += mac_us(256, H * W, 128 * 9)                            # convPa
    sp += mac_us(65, H * W, 256)                                 # convPb
    sp += 5 * ew_us(65, H * W)                                   # detector softmax
    Hf, Wf = H * 8, W * 8
    sp += taps_us(Hf, Wf, 243)                                   # simple_nms: 243 taps
    sp += ew_us(Hf, Wf)                                          # threshold
    sp += taps_us(Hf, Wf, 30)                                    # topk calibration iters
    sp += mac_us(256, H * W, 128 * 9)                            # convDa
    sp += mac_us(256, H * W, 256)                                # convDb
    sp += 3 * ew_us(256, H * W)                                  # dense L2
    sp += mac_us(256, N, 4)                                      # grid_sample (real MAC)
    sp += 3 * ew_us(256, N)                                      # per-kpt L2
    sp_total = 2 * sp
    # SuperGlue keypoint encoder x2
    enc = 0
    for cin, cout in [(3, 32), (32, 64), (64, 128), (128, 256)]:
        enc += mac_us(cout, N, cin)
    enc += ew_us(D, N)
    enc *= 2
    # GNN block (per-head core x4)
    head = mac_us(N, N, HD) + ew_us(N, N) + 5 * ew_us(N, N) + mac_us(N, HD, N)
    block = 3 * mac_us(D, N, D) + 4 * head + mac_us(D, N, D) + ew_us(D, N) \
        + mac_us(2 * D, N, 2 * D) + ew_us(2 * D, N) + mac_us(D, N, 2 * D) + ew_us(D, N)
    gnn = GNN_BLOCKS * block
    # Matching
    fin = 2 * mac_us(D, N, D)
    score = mac_us(N, N, D)
    scl = ew_us(N, N)
    if variant == "max_sinkhorn":
        sink_one = 2 * ew_us(NP1, NP1)                          # row: sum+scale
        sink_col_extra = ew_us(NP1, NP1)                        # transpose
        sink_one2 = 2 * ew_us(NP1, NP1)                         # col: sum+scale
        sink = T * (sink_one + sink_col_extra + sink_one2)
    elif variant == "transpose_free":
        sink = T * (5 * ew_us(NP1, NP1) + 0 + 5 * ew_us(NP1, NP1))
    else:
        sink = T * (5 * ew_us(NP1, NP1) + ew_us(NP1, NP1) + 5 * ew_us(NP1, NP1))
    if variant == "fixed_thresh":
        read = 2 * ew_us(NP1, NP1)
    else:
        read = 5 * ew_us(NP1, NP1)
    return sp_total + enc + gnn + fin + score + scl + sink + read


# ---------------- formula evaluator (for spreadsheet self-check) ------------
CELLRE = re.compile(r'\$?([A-Z]+)\$?([0-9]+)')
SUMRE = re.compile(r'SUM\(\$?([A-Z]+)\$?([0-9]+):\$?([A-Z]+)\$?([0-9]+)\)')


def evaluate(ws, addr):
    cache = {}
    def cell(a):
        if a in cache: return cache[a]
        cache[a] = 0.0
        v = ws[a].value
        if v is None:
            cache[a] = 0.0; return 0.0
        if isinstance(v, (int, float)):
            cache[a] = float(v); return cache[a]
        s = str(v)
        if not s.startswith('='):
            try: cache[a] = float(s)
            except ValueError: cache[a] = 0.0
            return cache[a]
        e = SUMRE.sub(lambda m: '(' + '+'.join(f'{m.group(1)}{i}'
                       for i in range(int(m.group(2)), int(m.group(4)) + 1)) + ')', s[1:])
        e = CELLRE.sub(lambda m: f'cell("{m.group(1)}{m.group(2)}")', e)
        try: cache[a] = eval(e, {'cell': cell})
        except Exception: cache[a] = float('nan')
        return cache[a]
    return cell(addr)


# ---------------- build all sheets ------------------------------------------
VARIANTS = [("original", "SuperGlue (original)"),
            ("max_sinkhorn", "SG max-Sinkhorn"),
            ("base2_softmax", "SG base2-softmax"),
            ("transpose_free", "SG transpose-free col"),
            ("fixed_thresh", "SG fixed-thresh match")]
gt_by_sheet = {}
for v, nm in VARIANTS:
    _, gt = build_sheet(v, nm)
    gt_by_sheet[nm] = (v, gt)

meas_ws, meas_gt, meas_total_cyc, meas_rates = build_measured_sheet()

wb.save(OUT)

# ---------------- validation -------------------------------------------------
print(f"\nwrote {OUT}\nsheets: {wb.sheetnames}\n")
print(f"{'sheet':30s}  {'PY model [us]':>14s}  {'XLS Z grand [us]':>16s}  {'FPS':>5s}  status")
print("-" * 80)
all_ok = True
for v, nm in VARIANTS:
    py = py_total(v)
    ws = wb[nm]; gt = gt_by_sheet[nm][1]
    xls = evaluate(ws, f"Z{gt}")
    fps = evaluate(ws, f"Z{gt + 1}")
    ok = abs(py - xls) / max(py, 1) < 1e-9
    all_ok &= ok
    print(f"{nm:30s}  {py:14.1f}  {xls:16.1f}  {fps:5.2f}  {'PASS' if ok else 'FAIL'}")
nm = "SG max-Sinkhorn (measured)"
ws = wb[nm]
xls_meas = evaluate(ws, f"Z{meas_gt}")
fps_meas = evaluate(ws, f"Z{meas_gt + 1}")
print(f"{nm:30s}  {meas_total_cyc/1000:14.1f}  {xls_meas:16.1f}  {fps_meas:5.2f}  "
      f"{'PASS' if abs(meas_total_cyc/1000 - xls_meas)/max(xls_meas,1)<1e-6 else 'FAIL'}")
print(f"\nmeasured rates (emulator):", {k: round(v, 3) for k, v in meas_rates.items()})
print("\n", "ALL PASS" if all_ok else "SOME FAILED")
