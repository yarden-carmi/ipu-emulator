#!/usr/bin/env python3
"""Build SuperGlue performance-model sheets in the MobileViT spreadsheet's style.

One sheet per variant: original SuperGlue + each of our 4 modifications isolated.
Cells use the same column layout and formula conventions as the MobileViT sheet
(Cycles = work/128 lanes, Time=Cycles/freq, BW = mem*8/Time). N=512 keypoints,
D=256 descriptor dim, 4 heads (head_dim 64), 18 GNN layers (self/cross) over a
pair of images, Sinkhorn T=100 on the (N+1)x(N+1) assignment matrix.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

SRC = "/root/.claude/uploads/8c029b26-5a4c-4fdb-96ec-f1feb856d256/fa8d8c24-IPU_2.xlsx"
OUT = "/home/user/ipu-emulator/IPU_2_SuperGlue.xlsx"

N, D, HEADS, HD, NP1, T = 512, 256, 4, 64, 513, 100
GNN_BLOCKS = 18 * 2          # 18 self/cross layers x 2 images
COLS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

wb = openpyxl.load_workbook(SRC)
mob = wb["MobileViT"]
HEADER = [mob.cell(1, c).value for c in range(1, 27)]   # A1..Z1


def make_sheet(name):
    ws = wb.create_sheet(name)
    for c in range(1, 27):
        ws.cell(1, c, HEADER[c - 1])
        ws.cell(1, c).font = copy(mob.cell(1, c).font)
    for col, w in [("A", 30), ("B", 22), ("C", 5), ("D", 6), ("E", 6), ("Y", 30), ("Z", 22)]:
        ws.column_dimensions[col].width = w
    ws["Y2"], ws["Z2"] = 1, 8          # freq GHz, Mem BW (same as MobileViT)
    return ws


# cycle-formula families (templates use {r} = row number) -----------------------
def cyc(kind, r):
    f = {
        "linear":  f"=C{r}*D{r}*G{r}/128",            # Cin*N*Cout
        "mm2":     f"=(C{r}*D{r}*G{r}/128)*2",         # GEMM (mul+add)
        "mmv":     f"=C{r}*D{r}*F{r}/128",             # attn @ V
        "scale":   f"=C{r}*D{r}/128",
        "softmax": f"=C{r}*D{r}*5/128",                # max,sub,exp,sum,div
        "resid":   f"=C{r}*D{r}/128",
        "reshape": "0",
        "norm5":   f"=C{r}*D{r}*5/128",                # log-domain logsumexp norm
        "norm2":   f"=C{r}*D{r}*2/128",                # plain sum+scale norm
        "transp":  f"=C{r}*D{r}/128",
        "pass1":   f"=C{r}*D{r}/128",                  # single-pass max/threshold
    }
    return f[kind]


def emit(ws, r, layer, typ, c, d, e, f, g, h, i, j, k, l, m, n, kind, note=None, ynote=None):
    vals = {"A": layer, "B": typ, "C": c, "D": d, "E": e, "F": f, "G": g, "H": h,
            "I": i, "J": j, "K": k, "L": l, "M": m, "N": n,
            "O": f"=C{r}*D{r}/1024", "P": f"=F{r}*G{r}/1024", "Q": 0,
            "R": f"=L{r}*M{r}/1024", "S": cyc(kind, r)}
    if kind == "reshape":
        vals["S"] = "0"
        vals["T"] = 0
        for cc in "UVWX":
            vals[cc] = 0
    else:
        vals["T"] = f"=(S{r}/$Y$2)/1000"
        vals["U"] = f"=((O{r}*8*1024)/$T{r})/1000"
        vals["V"] = f"=((P{r}*8*1024)/$T{r})/1000"
        vals["W"] = f"=((Q{r}*8*1024)/$T{r})/1000"
        vals["X"] = f"=((R{r}*8*1024)/$T{r})/1000"
    for col, v in vals.items():
        ws.cell(r, COLS.index(col) + 1, v)
    if ynote:
        ws.cell(r, 25, ynote)        # col Y label
    return r + 1


def build(ws, variant):
    """variant in {'original','max_sinkhorn','base2_softmax','transpose_free','fixed_thresh'}."""
    r = 2
    # ---- Keypoint encoder (per image; x2) ----  (Y2/Z2 hold freq/MemBW -- never label row 2)
    r = emit(ws, r, "Input (kpt+desc)", "Model input", D, N, 1, 0, 0, 0, 0, 0, 0, D, N, 1, "reshape")
    for cin, cout in [(3, 32), (32, 64), (64, 128), (128, 256)]:
        r = emit(ws, r, "MLP + BN + ReLU", f"KptEnc {cin}->{cout}", cin, N, cin, cin, cout, cout, cout, 0, 0, cout, N, cout, "linear")
    enc_total = r
    r = emit(ws, r, "Residual", "desc += enc(kpt)", D, N, 0, D, N, 0, 0, 0, 0, D, N, 0, "resid", ynote="2X (two images)")
    ws.cell(enc_total - 4, 26, f"=2*(SUM(T{enc_total-4}:T{enc_total}))")  # encoder total (col Z)
    ws.cell(enc_total - 4, 25, "Encoder total [us] ->")

    # ---- One GNN attention block (expanded; x GNN_BLOCKS) ----
    blk0 = r
    r = emit(ws, r, "Projection", "Q proj (D->D)", D, N, 0, D, D, D, 0, 0, 0, D, N, 0, "linear",
             ynote=f"GNN block (x{GNN_BLOCKS}=18 layers x 2 img)")
    r = emit(ws, r, "Projection", "K proj (D->D)", D, N, 0, D, D, D, 0, 0, 0, D, N, 0, "linear")
    r = emit(ws, r, "Projection", "V proj (D->D)", D, N, 0, D, D, D, 0, 0, 0, D, N, 0, "linear")
    r = emit(ws, r, "Reshape", "head reshape Q", D, N, 0, 0, 0, 0, 0, 0, 0, HD, N, HEADS, "reshape")
    r = emit(ws, r, "Reshape", "head reshape K", D, N, 0, 0, 0, 0, 0, 0, 0, HD, N, HEADS, "reshape")
    r = emit(ws, r, "Reshape", "head reshape V", D, N, 0, 0, 0, 0, 0, 0, 0, HD, N, HEADS, "reshape")
    r_qkt = r
    r = emit(ws, r, "Mat mul", "QKt (1 head)", HD, N, 0, HD, N, 0, 0, 0, 0, N, N, 0, "mm2", ynote="4X (heads)")
    r_sc = r
    r = emit(ws, r, "Scaling", "1/sqrt(d)", N, N, 0, 0, 0, 0, 0, 0, 0, N, N, 0, "scale")
    r_sm = r
    sm_kind = "softmax"
    r = emit(ws, r, "Softmax", "attn softmax" + (" (exp2)" if variant == "base2_softmax" else ""),
             N, N, 0, 0, 0, 0, 0, 0, 0, N, N, 0, sm_kind)
    r_qkv = r
    r = emit(ws, r, "Mat mul", "attn @ V (1 head)", N, N, 0, HD, N, 0, 0, 0, 0, N, HD, 0, "mmv")
    r = emit(ws, r, "Reshape", "concat heads", HD, N, HEADS, 0, 0, 0, 0, 0, 0, D, N, 0, "reshape",
             ynote="total 4 heads [us]")
    ws.cell(r - 1, 26, f"=4*(T{r_qkt}+T{r_sc}+T{r_sm}+T{r_qkv})")
    r = emit(ws, r, "Mat mul", "out proj (D->D)", D, N, 0, D, D, D, 0, 0, 0, D, N, 0, "linear")
    r = emit(ws, r, "Residual", "x += message", D, N, 0, D, N, 0, 0, 0, 0, D, N, 0, "resid")
    r = emit(ws, r, "Mat mul + ReLU", "merge MLP 2D->2D", 2 * D, N, 0, 2 * D, 2 * D, 2 * D, 0, 0, 0, 2 * D, N, 0, "linear")
    r = emit(ws, r, "Mat mul", "merge MLP 2D->D", 2 * D, N, 0, 2 * D, D, D, 0, 0, 0, D, N, 0, "linear")
    r = emit(ws, r, "Residual", "x += mlp", D, N, 0, D, N, 0, 0, 0, 0, D, N, 0, "resid")
    blk_end = r - 1
    # GNN total: sum of block T's (with the 4-head total cell) x GNN_BLOCKS
    ws.cell(blk0, 25, ws.cell(blk0, 25).value)
    ws.cell(blk0, 26, f"={GNN_BLOCKS}*(SUM(T{blk0}:T{blk_end})+Z{r_qkv+1}-T{r_qkt}-T{r_sc}-T{r_sm}-T{r_qkv})")
    ws.cell(blk0 + 1, 25, "GNN total [us] ->")
    ws.cell(blk0 + 1, 26, f"={GNN_BLOCKS}*(SUM(T{blk0}:T{blk_end})+Z{r_qkv+1}-T{r_qkt}-T{r_sc}-T{r_sm}-T{r_qkv})")

    # ---- Matching: final proj, score, Sinkhorn, readout ----
    r_final = r
    r = emit(ws, r, "Projection", "final proj (D->D)", D, N, 0, D, D, D, 0, 0, 0, D, N, 0, "linear",
             ynote="Matching (x2 final proj)")
    r_score = r
    r = emit(ws, r, "Mat mul", "score = A^T B /sqrt d", D, N, 0, D, N, 0, 0, 0, 0, N, N, 0, "mm2")
    r = emit(ws, r, "Reshape", "dustbin augment", NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, "reshape")

    # Sinkhorn iteration (row norm, col transpose, col norm) x T
    norm_kind = "norm2" if variant == "max_sinkhorn" else "norm5"
    rsk = r
    r = emit(ws, r, "Sinkhorn", "row normalize", NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, norm_kind,
             ynote=f"Sinkhorn x{T} iters")
    if variant == "transpose_free":
        r = emit(ws, r, "Sinkhorn", "col transpose (eliminated)", NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, "reshape")
    else:
        r = emit(ws, r, "Sinkhorn", "col transpose", NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, "transp")
    r = emit(ws, r, "Sinkhorn", "col normalize", NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, norm_kind)
    sk_end = r - 1
    ws.cell(rsk, 26, f"={T}*(SUM(T{rsk}:T{sk_end}))")
    ws.cell(rsk + 1, 25, f"Sinkhorn total (x{T}) [us] ->")
    ws.cell(rsk + 1, 26, f"={T}*(SUM(T{rsk}:T{sk_end}))")

    read_kind = "pass1" if variant == "fixed_thresh" else "softmax"
    read_lbl = "argmax + fixed threshold" if variant == "fixed_thresh" else "dual-softmax readout"
    r = emit(ws, r, "Matching", read_lbl, NP1, NP1, 0, 0, 0, 0, 0, 0, 0, NP1, NP1, 0, read_kind,
             ynote="Match readout")

    # ---- Grand total ----
    r_read = r - 1
    r += 1
    ws.cell(r, 1, "GRAND TOTAL [us]")
    ws.cell(r, 25, "variant: " + variant)
    ws.cell(r, 26, f"=Z{enc_total-4}+Z{blk0+1}+2*T{r_final}+T{r_score}+Z{rsk+1}+T{r_read}")
    return ws


for v, nm in [("original", "SuperGlue (original)"),
              ("max_sinkhorn", "SG max-Sinkhorn"),
              ("base2_softmax", "SG base2-softmax"),
              ("transpose_free", "SG transpose-free col"),
              ("fixed_thresh", "SG fixed-thresh match")]:
    build(make_sheet(nm), v)

wb.save(OUT)
print("wrote", OUT)
print("sheets:", wb.sheetnames)
