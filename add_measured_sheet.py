#!/usr/bin/env python3
"""Append a MEASURED max-Sinkhorn sheet: same SuperGlue op rows, but the Cycles
column holds cycles measured by actually running each kernel in the IPU emulator
(via superpoint_eval/measure_kernels.rates()), scaled to the SuperGlue sizes by
128-lane tiling. Op->kernel: mac=matmul inner loop, ew=elementwise MULT.EE,
softmax=softmax.asm, sinkrow=sinkhorn_iter.asm, sinkcol=sinkhorn_col.asm (both
already the max/non-log variant). Same MobileViT formatting; adds FPS.
"""
import os, sys, math
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

sys.path.insert(0, os.path.join(os.path.dirname(__file__),
                                "kernels/superpoint_superglue/superpoint_eval"))
from measure_kernels import rates

FILE = "/home/user/ipu-emulator/IPU_2_SuperGlue.xlsx"
N, D, HEADS, HD, NP1, T = 512, 256, 4, 64, 513, 100
GNN_BLOCKS = 18 * 2
COLS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

R = rates()
print("measured rates:", {k: round(v, 3) for k, v in R.items()})


def tiles(e):
    return math.ceil(e / 128)

def opcyc(kind, c, d, K):
    """Measured cycles for one instance of an op at the given size."""
    if kind == "mac":      # GEMM: out (c*d) elems, contraction K
        return tiles(c * d) * (R["mac_step"] * K + R["tile_ovh"])
    if kind == "ew":
        return tiles(c * d) * R["ew_tile"]
    if kind == "softmax":  # softmax over rows of length c, d rows
        return d * tiles(c) * R["softmax128"]
    if kind == "sinkrow":  # matrix d rows x c cols, row half-step
        return d * tiles(c) * R["sink_row"]
    if kind == "sinkcol":
        return d * tiles(c) * R["sink_col"]
    return 0.0


# (group, type/op, c, d, K, kind, instance_multiplier)
ROWS = [
    ("Input", "kpt+desc load", D, N, 0, "reshape", 1),
]
for cin, cout in [(3, 32), (32, 64), (64, 128), (128, 256)]:
    ROWS.append((f"KptEnc {cin}->{cout}", "MAC matmul", cout, N, cin, "mac", 2))
ROWS.append(("Residual desc+=enc", "add", D, N, 0, "ew", 2))
# GNN block x36
g = GNN_BLOCKS
ROWS += [
    ("Q proj D->D", "MAC matmul", D, N, D, "mac", g),
    ("K proj D->D", "MAC matmul", D, N, D, "mac", g),
    ("V proj D->D", "MAC matmul", D, N, D, "mac", g),
    ("QKt", "MAC matmul", N, N, HD, "mac", g * HEADS),
    ("scale 1/sqrt(d)", "multiply", N, N, 0, "ew", g * HEADS),
    ("softmax", "max,exp2,sum,mul", N, N, 0, "softmax", g * HEADS),
    ("attn @ V", "MAC matmul", N, HD, N, "mac", g * HEADS),
    ("out proj D->D", "MAC matmul", D, N, D, "mac", g),
    ("Residual x+=msg", "add", D, N, 0, "ew", g),
    ("merge MLP 2D->2D", "MAC matmul", 2 * D, N, 2 * D, "mac", g),
    ("merge MLP relu", "relu", 2 * D, N, 0, "ew", g),
    ("merge MLP 2D->D", "MAC matmul", D, N, 2 * D, "mac", g),
    ("Residual x+=mlp", "add", D, N, 0, "ew", g),
    ("final proj D->D", "MAC matmul", D, N, D, "mac", 2),
    ("score = A^T B", "MAC matmul", N, N, D, "mac", 1),
    ("scale 1/sqrt(d)", "multiply", N, N, 0, "ew", 1),
    ("Sinkhorn row norm (max)", "sum,scale", NP1, NP1, 0, "sinkrow", T),
    ("Sinkhorn col norm (max)", "ACC.MAX,scale", NP1, NP1, 0, "sinkcol", T),
    ("match readout", "argmax+thresh", NP1, NP1, 0, "ew", 1),
]

wb = openpyxl.load_workbook(FILE)
mob = wb["MobileViT"]
name = "SG max-Sinkhorn (measured)"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)


def cstyle(dst, src):
    dst.font = copy(src.font); dst.fill = copy(src.fill); dst.border = copy(src.border)
    dst.alignment = copy(src.alignment); dst.number_format = src.number_format


for c in range(1, 27):
    ws.cell(1, c, mob.cell(1, c).value); cstyle(ws.cell(1, c), mob.cell(1, c))
for i in range(1, 27):
    L = get_column_letter(i); ws.column_dimensions[L].width = mob.column_dimensions[L].width
ws.freeze_panes = "B2"
ws["Y2"], ws["Z2"] = 1, 8

r = 2
total_cyc = 0.0
for group, op, c, d, K, kind, mult in ROWS:
    one = opcyc(kind, c, d, K)
    cyc = round(one * mult)
    total_cyc += cyc
    vals = {"A": group, "B": op, "C": c, "D": d, "E": K, "F": 0, "G": K, "H": mult,
            "I": 0, "J": 0, "K": 0, "L": c, "M": d, "N": 0,
            "O": f"=C{r}*D{r}/1024", "P": 0, "Q": 0, "R": f"=L{r}*M{r}/1024",
            "S": cyc, "T": f"=(S{r}/$Y$2)/1000"}
    if kind == "reshape" or cyc == 0:
        vals["S"] = 0; vals["T"] = 0
        for cc in "UVWX": vals[cc] = 0
    else:
        for cc, src in [("U", "O"), ("V", "P"), ("W", "Q"), ("X", "R")]:
            vals[cc] = f"=(({src}{r}*8*1024)/$T{r})/1000"
    for col, v in vals.items():
        cstyle(ws.cell(r, COLS.index(col) + 1, v), mob.cell(5, COLS.index(col) + 1))
    ws.cell(r, 25, f"x{mult}" if mult > 1 else "");
    r += 1

# totals
gt = r
cstyle(ws.cell(gt, 1, "GRAND TOTAL [us] (measured)"), mob.cell(5, 1))
cstyle(ws.cell(gt, 25, "measured emulator cycles ->"), mob.cell(68, 25))
cstyle(ws.cell(gt, 26, f"=SUM(T2:T{r-1})"), mob.cell(68, 26))
fr = r + 1
cstyle(ws.cell(fr, 1, "FPS (image pair)"), mob.cell(5, 1))
cstyle(ws.cell(fr, 25, "frames / sec ->"), mob.cell(68, 25))
cstyle(ws.cell(fr, 26, f"=1000000/Z{gt}"), mob.cell(68, 26))

wb.save(FILE)
print(f"measured total = {total_cyc/1000:.1f} us  ->  {1e6/(total_cyc/1000):.2f} FPS")
print("sheets:", wb.sheetnames)
